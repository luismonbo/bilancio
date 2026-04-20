#!/usr/bin/env python
"""inspect-xlsx.py — print the first 20 rows of an XLSX file to diagnose parser detection.

Usage:
    uv run python scripts/inspect-xlsx.py <path-to-file.xlsx>
"""

import sys
from pathlib import Path

import openpyxl


def dump_sheet(ws: object, label: str, max_row: int = 20) -> None:
    dims = getattr(ws, "dimensions", "n/a")
    print(f"  Sheet: {label!r}  (dimensions: {dims})")
    found_any = False
    for row_idx, row in enumerate(ws.iter_rows(values_only=True, max_row=max_row)):
        nonempty = {i: v for i, v in enumerate(row) if v is not None and v != ""}
        if nonempty:
            found_any = True
            print(f"    Row {row_idx + 1:>2} (0-based {row_idx}): {nonempty}")
    if not found_any:
        print("    (all rows empty in first", max_row, "rows)")
    print()


def main() -> None:
    if len(sys.argv) != 2:
        print("Usage: uv run python scripts/inspect-xlsx.py <path-to-file.xlsx>")
        sys.exit(1)

    path = Path(sys.argv[1])
    if not path.exists():
        print(f"File not found: {path}")
        sys.exit(1)

    print(f"File: {path}")
    print()

    # --- Pass 1: read_only=True, data_only=True (same as the parser) ---
    print("=== Pass 1: read_only=True, data_only=True (same as parser) ===")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    print(f"  Sheets: {wb.sheetnames}  |  active: {wb.active.title if wb.active else 'none'}")
    print()
    for name in wb.sheetnames:
        dump_sheet(wb[name], name)

    # --- Pass 2: read_only=False (loads cached formula results + styles) ---
    print("=== Pass 2: read_only=False, data_only=True ===")
    wb2 = openpyxl.load_workbook(path, read_only=False, data_only=True)
    print(f"  Sheets: {wb2.sheetnames}  |  active: {wb2.active.title if wb2.active else 'none'}")
    print()
    for name in wb2.sheetnames:
        dump_sheet(wb2[name], name)

    print("Parser expects (on whichever sheet has data):")
    print("  Row  1 (index 0), col 1 == 'LISTA MOVIMENTI'")
    print("  Row 15 (index 14), cols 1-6 == "
          "('Data contabile', 'Data valuta', 'Tipologia', 'Entrate', 'Uscite', 'Divisa')")


if __name__ == "__main__":
    main()
