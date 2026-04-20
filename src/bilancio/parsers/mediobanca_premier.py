"""Parser for Mediobanca Premier XLSX exports.

File structure
--------------
Rows 1–14   Header block (account metadata, filter info).
Row  15     Column headers: Data contabile | Data valuta | Tipologia | Entrate | Uscite | Divisa
Rows 16–N   Transaction data (N varies by export period).
Last 3 rows Footer: blank row, totals header, totals values.

Column layout (1-based in spreadsheet, 0-based in openpyxl tuple):
  0  – always None (blank column A)
  1  – Data contabile (booking date, DD/MM/YYYY string; may be empty string)
  2  – Data valuta    (value date, DD/MM/YYYY string; always present)
  3  – Tipologia      (raw description, encodes transaction type + merchant info)
  4  – Entrate        (inflow as positive float, or empty string)
  5  – Uscite         (outflow as negative float, or empty string)
  6  – Divisa         (currency, e.g. 'EUR')

Known transaction type prefixes (from Tipologia):
  Pagam. POS      – card payment at POS terminal
  Addebito SDD    – SEPA direct debit
  Bonif. v/fav.   – incoming bank transfer
  Bancomat        – ATM cash withdrawal
  Disposizione    – outgoing bank transfer
  Stipendio       – salary credit
  Addebito canone – bank account fee
  POS             – direct POS (compact format, no space before city)
  Imposta bollo   – stamp duty tax
"""

import hashlib
import re
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

from bilancio.parsers.base import BankParser, ParsedTransaction

# Spreadsheet layout constants (0-based tuple indices)
_COL_BOOKING_DATE = 1
_COL_VALUE_DATE = 2
_COL_TIPOLOGIA = 3
_COL_ENTRATE = 4
_COL_USCITE = 5
_COL_CURRENCY = 6

_HEADER_ROW_IDX = 14   # 0-based; spreadsheet row 15 with column names
_DATA_START_IDX = 15   # 0-based; spreadsheet row 16, first data row

_DATE_FMT = "%d/%m/%Y"

# Detection markers
_MARKER_CELL_COL = 1   # column B (0-based)
_MARKER_VALUE = "LISTA MOVIMENTI"
_EXPECTED_HEADERS = ("Data contabile", "Data valuta", "Tipologia", "Entrate", "Uscite", "Divisa")


# ---------------------------------------------------------------------------
# Merchant extraction
# ---------------------------------------------------------------------------

def _extract_type_and_merchant(tipologia: str) -> tuple[str, str | None]:
    """Return (transaction_type, merchant_clean) from a raw Tipologia string."""

    if tipologia.startswith("Pagam. POS"):
        # e.g. "Pagam. POS - PAGAMENTO POS 25,00 EUR DEL 05.04.2026 A (ITA) TRENITALIA - PT WL   CARTA ****"
        m = re.search(r"\([A-Z ]{2,3}\)\s+(.+?)\s{2,}CARTA", tipologia)
        merchant = m.group(1).strip() if m else None
        return "Pagam. POS", merchant

    if tipologia.startswith("Bancomat"):
        # e.g. "Bancomat - PRELIEVO E/C 100,00 EUR DEL 25.02.2026 A (ITA) POSTE ITALIANE   CARTA ****"
        m = re.search(r"\([A-Z ]{2,3}\)\s+(.+?)\s{2,}CARTA", tipologia)
        merchant = m.group(1).strip() if m else None
        return "Bancomat", merchant

    if tipologia.startswith("Addebito SDD"):
        # e.g. "Addebito SDD - ILIAD               -..."
        # e.g. "Addebito SDD - PayPal Europe S.a.r.-1049427834310/PAYPAL..."
        after = tipologia.removeprefix("Addebito SDD - ")
        m = re.match(r"(.+?)\s{3,}", after)
        merchant = m.group(1).strip() if m else after.split("-")[0].strip()
        return "Addebito SDD", merchant

    if tipologia.startswith("Bonif. v/fav."):
        # e.g. "Bonif. v/fav. - RIF:216378667ORD. PAYPAL INST INSTANT TRANSFER"
        # e.g. "Bonif. v/fav. - RIF:213807679ORD. IMMOBILIARE RIVA RENO S.R.L. Rimborso.../SEPASCT/"
        m = re.search(r"(?:ORD\.|BEN\.)\s+(.+?)(?:/SEPASCT/|$)", tipologia)
        merchant = m.group(1).strip() if m else None
        return "Bonif. v/fav.", merchant

    if tipologia.startswith("Disposizione"):
        # e.g. "Disposizione - RIF:214625531BEN. Immobiliare Riva Di Reno S.R.L Affitto Feb 2026"
        m = re.search(r"BEN\.\s+(.+)", tipologia)
        merchant = m.group(1).strip() if m else None
        return "Disposizione", merchant

    if tipologia.startswith("Stipendio"):
        # e.g. "Stipendio - RIF:216722932ORD. DATA SCIENCE OPERATIONS S.R.L. ACCREDITO..."
        m = re.search(r"ORD\.\s+(.+?)\s+ACCREDITO", tipologia)
        merchant = m.group(1).strip() if m else None
        return "Stipendio", merchant

    if tipologia.startswith("Addebito canone"):
        return "Addebito canone", "Mediobanca Premier"

    if tipologia.startswith("Imposta bollo"):
        return "Imposta bollo", "Mediobanca Premier"

    if tipologia.startswith("POS-"):
        # Compact format: "POS-<merchant><city>" with no separator.
        # Return everything after "POS-" as-is; the rules engine can refine later.
        return "POS", tipologia.removeprefix("POS-")

    # Unknown prefix — preserve raw as type, no merchant extraction
    return tipologia.split("-")[0].strip(), None


# ---------------------------------------------------------------------------
# Hash
# ---------------------------------------------------------------------------

def _make_hash(value_date: datetime, amount: float, description_raw: str) -> str:
    """Stable SHA-256 hash for dedup. Does not include account_id (uniqueness
    is enforced by the (account_id, hash) DB constraint)."""
    content = f"{value_date.isoformat()}|{amount}|{description_raw}"
    return hashlib.sha256(content.encode()).hexdigest()


# ---------------------------------------------------------------------------
# Date parsing
# ---------------------------------------------------------------------------

def _parse_date(raw: object) -> datetime | None:
    if not raw or raw == "":
        return None
    if isinstance(raw, datetime):
        return raw.replace(tzinfo=timezone.utc)
    try:
        return datetime.strptime(str(raw), _DATE_FMT).replace(tzinfo=timezone.utc)
    except ValueError:
        return None


# ---------------------------------------------------------------------------
# Parser
# ---------------------------------------------------------------------------

class MediobancaPremierParser:
    bank_name: str = "Mediobanca Premier"

    def detect(self, file_path: Path) -> bool:
        """Return True if the file looks like a Mediobanca Premier XLSX export."""
        if not file_path.exists():
            return False
        if file_path.suffix.lower() != ".xlsx":
            return False
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            if ws is None:
                return False
            rows = list(ws.iter_rows(values_only=True, max_row=16))
            # Row 1 (index 0): col B must contain the marker
            if len(rows) < 16:
                return False
            marker_ok = rows[0][_MARKER_CELL_COL] == _MARKER_VALUE
            # Row 15 (index 14): column headers
            header_row = rows[_HEADER_ROW_IDX]
            headers_ok = tuple(header_row[1:7]) == _EXPECTED_HEADERS
            return bool(marker_ok and headers_ok)
        except Exception:
            return False

    def parse(self, file_path: Path, account_id: int) -> list[ParsedTransaction]:
        """Parse all transactions from the file. Skips header and footer rows."""
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        if ws is None:
            return []

        results: list[ParsedTransaction] = []

        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx < _DATA_START_IDX:
                continue

            tipologia = row[_COL_TIPOLOGIA]
            # Stop at footer rows (empty tipologia or totals labels)
            if not tipologia or str(tipologia).startswith("Totale"):
                continue

            value_date = _parse_date(row[_COL_VALUE_DATE])
            if value_date is None:
                continue  # malformed row — skip

            booking_date = _parse_date(row[_COL_BOOKING_DATE])

            entrate = row[_COL_ENTRATE]
            uscite = row[_COL_USCITE]
            amount = float(entrate) if entrate not in (None, "") else float(uscite)  # type: ignore[arg-type]

            currency = str(row[_COL_CURRENCY]) if row[_COL_CURRENCY] else "EUR"
            description_raw = str(tipologia)
            transaction_type, merchant_clean = _extract_type_and_merchant(description_raw)

            results.append(
                ParsedTransaction(
                    account_id=account_id,
                    booking_date=booking_date,
                    value_date=value_date,
                    amount=amount,
                    currency=currency,
                    transaction_type=transaction_type,
                    description_raw=description_raw,
                    merchant_clean=merchant_clean,
                    source_file=str(file_path),
                    source_row=row_idx + 1,  # 1-based spreadsheet row number
                    hash=_make_hash(value_date, amount, description_raw),
                )
            )

        return results


# Ensure the class satisfies the Protocol at import time
_: BankParser = MediobancaPremierParser()  # type: ignore[assignment]
